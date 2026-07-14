"""Parseo e importación de televisores desde CSV/XLSX.

Columnas reconocidas (encabezado, insensible a mayúsculas):
    - mac_address   (obligatoria, identifica el televisor)
    - serial_number (opcional)
    - numero_credito(opcional, solo dígitos)
"""
from __future__ import annotations

import csv
import io

from django.core.exceptions import ValidationError
from django.db import transaction

from televisores.models import Televisor

COLUMNAS = ('mac_address', 'serial_number', 'numero_credito')

# Filas por sentencia en bulk_create/bulk_update. Postgres tiene un tope de
# 65535 parámetros por consulta; 500 filas × pocas columnas queda muy por debajo
# y mantiene acotado el uso de memoria.
LOTE = 500


def _normalizar_encabezados(headers: list[str]) -> dict[int, str]:
    """Mapea índice de columna -> nombre de campo reconocido."""
    mapa: dict[int, str] = {}
    for i, h in enumerate(headers):
        clave = (h or '').strip().lower().replace(' ', '_')
        if clave in COLUMNAS:
            mapa[i] = clave
    return mapa


def _filas_csv(data: bytes) -> list[list[str]]:
    texto = data.decode('utf-8-sig', errors='replace')
    return [row for row in csv.reader(io.StringIO(texto))]


def _filas_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    filas: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        filas.append(['' if c is None else str(c) for c in row])
    wb.close()
    return filas


def leer_filas(nombre: str, data: bytes) -> list[list[str]]:
    """Lee un CSV/XLSX y devuelve sus filas como listas de strings."""
    nombre = (nombre or '').lower()
    if nombre.endswith(('.xlsx', '.xls')):
        return _filas_xlsx(data)
    return _filas_csv(data)


def importar_televisores(nombre: str, data: bytes, empresa) -> dict:
    """Crea/actualiza televisores de `empresa` desde el contenido de un archivo.

    Una consulta para leer los existentes y dos escrituras por lotes, en vez de
    un get_or_create + save por fila: el coste deja de crecer con el número de
    round-trips a la base de datos.

    Multi-tenant: solo se tocan los televisores de `empresa`. Una MAC (o un
    serial) que ya exista en OTRA empresa no se actualiza ni se duplica — se
    reporta como error de la fila. Sin esto, subir un archivo con la MAC de un
    equipo ajeno lo sobrescribiría.

    Devuelve {'creados', 'actualizados', 'errores': [str, ...]}.
    """
    filas = leer_filas(nombre, data)

    if not filas:
        return {'creados': 0, 'actualizados': 0, 'errores': ['El archivo está vacío.']}

    mapa = _normalizar_encabezados(filas[0])
    if 'mac_address' not in mapa.values():
        return {
            'creados': 0,
            'actualizados': 0,
            'errores': ['Falta la columna obligatoria "mac_address".'],
        }

    campos = set(mapa.values())
    errores: list[str] = []

    # Deduplica dentro del archivo conservando el orden de aparición: si una MAC
    # se repite, gana la última fila (mismo criterio que el get_or_create+save
    # secuencial que había antes, pero sin ir a la base de datos por cada una).
    orden: list[str] = []
    deseado: dict[str, dict] = {}

    for n, fila in enumerate(filas[1:], start=2):
        valores = {campo: (fila[i].strip() if i < len(fila) else '') for i, campo in mapa.items()}
        mac = valores.get('mac_address', '').upper()
        if not mac:
            continue  # fila vacía
        if len(mac) > 50:
            # bulk_create aborta el lote entero ante un error de la base de
            # datos, así que esto se valida aquí y no en el INSERT.
            errores.append(f'Fila {n} ({mac[:20]}…): la MAC supera los 50 caracteres.')
            continue
        if mac not in deseado:
            orden.append(mac)
        deseado[mac] = {'fila': n, 'valores': valores}

    if not orden:
        return {'creados': 0, 'actualizados': 0, 'errores': errores}

    # Una sola consulta por TODAS las empresas: hay que distinguir "ya lo tengo"
    # (actualizar) de "es de otra empresa" (rechazar), y ambos casos salen de la
    # misma búsqueda por MAC.
    encontrados = {
        tv.mac_address.upper(): tv
        for tv in Televisor.objects.filter(mac_address__in=orden)
    }
    existentes = {
        mac: tv for mac, tv in encontrados.items() if tv.empresa_id == empresa.pk
    }
    ajenos = {mac for mac, tv in encontrados.items() if tv.empresa_id != empresa.pk}

    # Seriales ya tomados en el sistema (por cualquier empresa, incluida esta con
    # otro equipo). Sin este control, el índice único haría fallar el bulk_create
    # entero en vez de rechazar solo la fila conflictiva.
    seriales_pedidos = {
        deseado[mac]['valores'].get('serial_number', '').strip()
        for mac in orden
    } - {''}
    seriales_tomados = {
        s.upper(): tv_mac
        for s, tv_mac in Televisor.objects.filter(
            serial_number__in=seriales_pedidos
        ).values_list('serial_number', 'mac_address')
    }

    nuevos: list[Televisor] = []
    actualizar: list[Televisor] = []
    vistos_serial: set[str] = set()

    for mac in orden:
        entrada = deseado[mac]
        valores = entrada['valores']

        if mac in ajenos:
            errores.append(
                f'Fila {entrada["fila"]} ({mac}): esta MAC ya está registrada en '
                'el sistema y no pertenece a tu empresa.'
            )
            continue

        tv = existentes.get(mac) or Televisor(mac_address=mac, empresa=empresa)

        serial = (valores.get('serial_number') or '').strip()
        if serial:
            dueño = seriales_tomados.get(serial.upper())
            # `dueño != mac`: que el serial ya esté en ESTE equipo no es conflicto.
            if (dueño and dueño.upper() != mac) or serial.upper() in vistos_serial:
                errores.append(
                    f'Fila {entrada["fila"]} ({mac}): el serial "{serial}" ya está '
                    'registrado en otro televisor.'
                )
                continue
            vistos_serial.add(serial.upper())

        if 'serial_number' in valores:
            tv.serial_number = valores['serial_number']
        if 'numero_credito' in valores:
            tv.numero_credito = valores['numero_credito']

        try:
            # validate_unique=False: la unicidad de la MAC ya la garantiza el
            # dedupe de arriba más el índice único; comprobarla aquí costaría un
            # SELECT por fila, que es justo lo que se quiere evitar. La empresa se
            # excluye por lo mismo: validarla dispararía un SELECT por fila para
            # comprobar que existe, y viene de una instancia ya cargada.
            tv.full_clean(exclude=['mac_address', 'empresa'], validate_unique=False)
        except ValidationError as exc:
            detalle = '; '.join(f'{k}: {" ".join(v)}' for k, v in exc.message_dict.items())
            errores.append(f'Fila {entrada["fila"]} ({mac}): {detalle}')
            continue

        (actualizar if tv.pk else nuevos).append(tv)

    # Los campos que Televisor.save() normaliza hay que normalizarlos a mano:
    # bulk_create y bulk_update no llaman a save().
    for tv in (*nuevos, *actualizar):
        tv.mac_address = tv.mac_address.strip().upper()
        tv.serial_number = (tv.serial_number or '').strip()
        tv.eui64 = (tv.eui64 or '').strip().upper()

    campos_editables = [c for c in ('serial_number', 'numero_credito') if c in campos]

    with transaction.atomic():
        if nuevos:
            Televisor.objects.bulk_create(nuevos, batch_size=LOTE)
        if actualizar and campos_editables:
            Televisor.objects.bulk_update(actualizar, campos_editables, batch_size=LOTE)

    return {'creados': len(nuevos), 'actualizados': len(actualizar), 'errores': errores}
