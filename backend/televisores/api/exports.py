"""Exportaciones a Excel (.xlsx) — sincronizaciones y códigos pin.

Mismo estilo que whaletv: encabezado con el color de marca (F6186A).
"""
from __future__ import annotations

import io

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from televisores.models import BulkSyncItem, BulkSyncJob, PinCodeUsado, SyncJob

from .filtros import filtrar_por_fecha, filtrar_sincronizaciones
from .registros import nombre_usuario

XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)


def _estilar_encabezado(ws):
    fill = PatternFill('solid', fgColor='F6186A')
    font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')


def _respuesta_xlsx(wb: Workbook, filename: str) -> HttpResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type=XLSX_CONTENT_TYPE)
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _resultado_syncjob(estado: str) -> str:
    return {'terminado': 'Aplicado', 'error': 'Error'}.get(estado, 'En proceso')


def _resultado_item(estado: str) -> str:
    return {'ok': 'Aplicado', 'error': 'Error'}.get(estado, 'Pendiente')


def _nombre_archivo(base: str, televisor=None, desde=None, hasta=None) -> str:
    """`base[_MAC][_rango].xlsx`. Los ':' de la MAC no valen en un nombre de
    archivo en Windows, así que se cambian por guiones."""
    partes = [base]
    if televisor is not None:
        partes.append(televisor.mac_address.replace(':', '-'))
    if desde or hasta:
        partes.append(f'{desde or "inicio"}_a_{hasta or "hoy"}')
    return '_'.join(partes) + '.xlsx'


def exportar_sincronizaciones(
    desde: str | None = None,
    hasta: str | None = None,
    televisor=None,
) -> HttpResponse:
    """Historial de cambios de estado (individuales + masivos), acotado al rango.

    Mismas filas que muestra /sincronizaciones: el filtro de la tabla y el del
    Excel salen de la misma función. Con `televisor` se limita a ese equipo,
    que es lo que exporta /televisores/<id>/sincronizaciones.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sincronizaciones'
    ws.append([
        'Fecha', 'Número de serial', 'Dirección MAC', 'Usuario',
        'Acción', 'Resultado', 'Tipo', 'Mensaje',
    ])
    _estilar_encabezado(ws)

    syncs = SyncJob.objects.all()
    # Solo los lotes de sincronización: los de validación masiva no son cambios
    # de estado, y la tabla de /sincronizaciones tampoco los muestra.
    items = BulkSyncItem.objects.filter(job__modo=BulkSyncJob.SYNC)
    if televisor is not None:
        syncs = syncs.filter(televisor=televisor)
        items = items.filter(televisor=televisor)

    syncs, items = filtrar_sincronizaciones(syncs, items, desde, hasta)

    filas = []
    # values_list evita instanciar los modelos; la MAC/serial del individual se
    # sacan por la relación, y el nombre del usuario se arma en SQL.
    for creado, serial, mac, usuario, inhabilitar, estado, error in (
        syncs.annotate(usuario_nombre=nombre_usuario('usuario')).values_list(
            'creado',
            'televisor__serial_number',
            'televisor__mac_address',
            'usuario_nombre',
            'inhabilitar',
            'estado',
            'error',
        ).iterator(chunk_size=2000)
    ):
        filas.append((
            creado,
            serial or '—',
            mac or '—',
            usuario or '—',
            'Inhabilitar' if inhabilitar else 'Habilitar',
            _resultado_syncjob(estado),
            'Individual',
            error or '',
        ))
    for creado, serial, mac, usuario, inhabilitar, estado, mensaje in (
        items.annotate(usuario_nombre=nombre_usuario('job__usuario')).values_list(
            'job__creado',
            'televisor__serial_number',
            'mac_address',
            'usuario_nombre',
            'inhabilitar',
            'estado',
            'mensaje',
        ).iterator(chunk_size=2000)
    ):
        filas.append((
            creado,
            serial or '—',
            mac,
            usuario or '—',
            'Inhabilitar' if inhabilitar else 'Habilitar',
            _resultado_item(estado),
            'Masivo',
            mensaje or '',
        ))

    filas.sort(key=lambda f: f[0], reverse=True)
    for f in filas:
        ws.append([
            timezone.localtime(f[0]).strftime('%d/%m/%Y %H:%M'),
            f[1], f[2], f[3], f[4], f[5], f[6], f[7],
        ])

    for col, ancho in zip('ABCDEFGH', (18, 20, 20, 24, 14, 12, 12, 40)):
        ws.column_dimensions[col].width = ancho

    return _respuesta_xlsx(
        wb, _nombre_archivo('sincronizaciones', televisor, desde, hasta)
    )


def _txt_bool(valor: bool | None, si: str, no: str) -> str:
    if valor is None:
        return '—'
    return si if valor else no


def exportar_bulk_job(job: BulkSyncJob) -> HttpResponse:
    """Exporta el detalle de UN lote (sincronización o validación masiva)."""
    es_validacion = job.modo == BulkSyncJob.VALIDACION
    wb = Workbook()
    ws = wb.active
    ws.title = 'Validación' if es_validacion else 'Sincronización'

    items = job.items.select_related('televisor')
    if es_validacion:
        ws.append(['Dirección MAC', 'Serial', 'Portal', 'App', 'Coincide', 'Mensaje'])
        _estilar_encabezado(ws)
        for it in items:
            ws.append([
                it.mac_address,
                it.televisor.serial_number if it.televisor else '',
                _txt_bool(it.remoto_inhabilitado, 'Inhabilitado', 'Habilitado'),
                _txt_bool(it.local_inhabilitado, 'Inhabilitado', 'Habilitado'),
                _txt_bool(it.coincide, 'Sí', 'No'),
                it.mensaje or '',
            ])
        anchos = zip('ABCDEF', (20, 20, 16, 16, 12, 40))
    else:
        ws.append(['Dirección MAC', 'Serial', 'Acción', 'Resultado', 'Mensaje'])
        _estilar_encabezado(ws)
        for it in items:
            ws.append([
                it.mac_address,
                it.televisor.serial_number if it.televisor else '',
                'Inhabilitar' if it.inhabilitar else 'Habilitar',
                _resultado_item(it.estado),
                it.mensaje or '',
            ])
        anchos = zip('ABCDE', (20, 20, 14, 12, 40))

    for col, ancho in anchos:
        ws.column_dimensions[col].width = ancho

    prefijo = 'validacion_masiva' if es_validacion else 'sincronizacion_masiva'
    return _respuesta_xlsx(wb, f'{prefijo}_{job.pk}.xlsx')


def plantilla_televisores() -> HttpResponse:
    """Plantilla Excel para enrolar televisores."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Televisores'
    ws.append(['mac_address', 'serial_number', 'numero_credito'])
    _estilar_encabezado(ws)
    ws.append(['B4:04:29:7E:3A:AA', 'B4:04:29:7E:3A:AA', '1234567890'])
    for col, ancho in zip('ABC', (22, 22, 18)):
        ws.column_dimensions[col].width = ancho
    return _respuesta_xlsx(wb, 'plantilla_televisores.xlsx')


def plantilla_estados() -> HttpResponse:
    """Plantilla Excel para enrolar estados (habilitado/inhabilitado)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Estados'
    ws.append(['mac_address', 'estado', 'serial_number'])
    _estilar_encabezado(ws)
    ws.append(['B4:04:29:7E:3A:AA', 'inhabilitado', 'B4:04:29:7E:3A:AA'])
    ws.append(['B4:04:29:7E:3A:BB', 'habilitado', 'B4:04:29:7E:3A:BB'])
    for col, ancho in zip('ABC', (22, 16, 22)):
        ws.column_dimensions[col].width = ancho
    return _respuesta_xlsx(wb, 'plantilla_estados.xlsx')


def exportar_pincodes(
    desde: str | None = None,
    hasta: str | None = None,
    televisor=None,
) -> HttpResponse:
    """Bitácora de Códigos Pin usados, acotada al mismo rango que ve el usuario.

    Lee de la base de datos, no del portal. La versión anterior consultaba la
    Device Lock API una vez por televisor: además de tardar minutos, devolvía
    los pines *disponibles* en el portal en vez de los ya *usados* en la app,
    que es lo que lista /pincodes. Con `televisor` se limita a ese equipo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Códigos Pin'
    ws.append([
        'Fecha', 'Número de serial', 'Dirección MAC', 'Código de Acceso', 'Código Pin',
    ])
    _estilar_encabezado(ws)

    qs = PinCodeUsado.objects.all()
    if televisor is not None:
        qs = qs.filter(televisor=televisor)

    filas = (
        filtrar_por_fecha(qs, desde, hasta)
        .order_by('-creado')
        # values_list + iterator: no construye instancias del modelo ni carga la
        # bitácora entera en memoria de golpe. El serial se saca por la relación.
        .values_list(
            'creado', 'televisor__serial_number', 'mac_address', 'passcode', 'pin_code'
        )
        .iterator(chunk_size=2000)
    )
    for creado, serial, mac, passcode, pin in filas:
        ws.append([
            timezone.localtime(creado).strftime('%d/%m/%Y %H:%M'),
            serial or '—',
            mac,
            passcode,
            pin,
        ])

    for col, ancho in zip('ABCDE', (18, 20, 20, 18, 16)):
        ws.column_dimensions[col].width = ancho

    return _respuesta_xlsx(wb, _nombre_archivo('pincodes', televisor, desde, hasta))
